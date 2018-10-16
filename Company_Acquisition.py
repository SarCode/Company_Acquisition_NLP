import openpyxl

#loading corpus
book = openpyxl.load_workbook('Corpus.xlsx')

sheet = book.active

id=[]
title=[]

for i in range (2,902):
    temp_i='A'+str(i)
    temp_title_i='B'+str(i)
    temp_title_sheet=sheet[temp_title_i]
    temp_i_sheet=sheet[temp_i]
    temp1=temp_i_sheet.value
    temp2=temp_title_sheet.value
    id.append(temp1)
    title.append(temp2)
    
    
##loading training data
    
book = openpyxl.load_workbook('MnA_Training.xlsx')

sheet = book.active

id_train=[]
acq_train=[]
tar_train=[]
for i in range (2,502):
    temp_i='A'+str(i)
    temp_i_sheet=sheet[temp_i]
    temp1=temp_i_sheet.value
    id_train.append(temp1)
    #acq
    temp_i_acq='B'+str(i)
    temp_i_acq_sheet=sheet[temp_i_acq]
    temp2=temp_i_acq_sheet.value
    acq_train.append(temp2)
    #tar
    temp_i_tar='C'+str(i)
    temp_i_tar_sheet=sheet[temp_i_tar]
    temp3=temp_i_tar_sheet.value
    tar_train.append(temp3)
    
    
##finding corresponding corpus and train data
c=0
corpus_train=[]
for i in range (0,500):
    temp=id.index(id_train[i])
    print(id[temp])
    print(title[temp])
    corpus_train.append(title[temp])
    
    
from isc_tokenizer import Tokenizer
tk = Tokenizer(lang='en')

from isc_tagger import Tagger
tagger = Tagger(lang='eng')

from __future__ import unicode_literals
from isc_parser import Parser
parser = Parser(lang='eng')


from nltk.stem.porter import PorterStemmer
ps=PorterStemmer()


tokenized=(tk.tokenize(corpus_train[10]))
print(tokenized)

print(tagger.tag(corpus_train[10].split()))


x=tagger.tag(corpus_train[10].split())
print(type(x))

print(tagger.tag(corpus_train[35].split()))
text=corpus_train[0]
text = text.split()

len(corpus_train)

"----------------------------BUY-------------------------------------------------------------------"            

temp_left=[]
temp_right=[]
left_tag=[]
right_tag=[]

tag=[]
prop_n_l=[]
prop_n_r=[]
corp_r=[]
corp_l=[]
buy_id=[]
buy_title=[]
check_buy=[]
for i in range(0,len(corpus_train)):
    temp = corpus_train[i]
    temp_token=tk.tokenize(temp)
    for j in range(0,len(temp_token)):
        if 'buy' in temp_token[j]:
            temp_left=temp_token[0:j]
            temp_right=temp_token[j+1:]
            left_tag=(tagger.tag(temp_left))
            right_tag=(tagger.tag(temp_right))
            print(left_tag)
            print(len(left_tag))
            for k in range(0,len(left_tag)):
                tag_l=left_tag[k]
                print(tag_l)
                if tag_l[1] == 'PROPN':
                    prop_n_l.append(tag_l[0])
            for l in range(0,len(right_tag)):
                tag_r=right_tag[l]
                print(tag_r)
                if tag_r[1] == 'PROPN':
                    prop_n_r.append(tag_r[0])
                    prop_n_r
            corp_l.append(prop_n_l)
            corp_r.append(prop_n_r)
            prop_n_l=[]
            buy_title.append(corpus_train[i])
            prop_n_r=[]
            buy_id.append(i)
            check_buy.append(id_train[i])
            


"----------------------------ACQUIRE--------------------------------------------------------------------"            

temp_left=[]
temp_right=[]
left_tag=[]
right_tag=[]

tag=[]
prop_n_l=[]
prop_n_r=[]
corp_r_acq=[]
corp_l_acq=[]
acq_id=[]
acq_title=[]
check_acq=[]
for i in range(0,len(corpus_train)):
    temp = corpus_train[i]
    temp_token=tk.tokenize(temp)
    for j in range(0,len(temp_token)):
        if 'acquir' in temp_token[j]:
            temp_left=temp_token[0:j]
            temp_right=temp_token[j+1:]
            left_tag=(tagger.tag(temp_left))
            right_tag=(tagger.tag(temp_right))
            print(left_tag)
            print(len(left_tag))
            for k in range(0,len(left_tag)):
                tag_l=left_tag[k]
                print(tag_l)
                if tag_l[1] == 'PROPN':
                    prop_n_l.append(tag_l[0])
            for l in range(0,len(right_tag)):
                tag_r=right_tag[l]
                print(tag_r)
                if tag_r[1] == 'PROPN':
                    prop_n_r.append(tag_r[0])
                    prop_n_r
            corp_l_acq.append(prop_n_l)
            corp_r_acq.append(prop_n_r)
            prop_n_l=[]
            acq_title.append(corpus_train[i])
            prop_n_r=[]
            acq_id.append(i)
            check_acq.append(id_train[i])
            
            
"----------------------------ACQUISITION--------------------------------------------------------------------"            

temp_left=[]
temp_right=[]
left_tag=[]
right_tag=[]

tag=[]
prop_n_l=[]
prop_n_r=[]
corp_r_acquisition=[]
corp_l_acquisition=[]
acquisition_id=[]
acquisition_title=[]
check_acquisition=[]
m=0
checker=0
try:
    for i in range(0,len(corpus_train)):
        temp = corpus_train[i]
        temp_token=tk.tokenize(temp)
        for j in range(0,len(temp_token)):
            if 'acquisition' in temp_token[j]:
                if j != (len(temp_token)-1):
                    temp_left=temp_token[0:j]
                    temp_right=temp_token[j+1:]
                    left_tag=(tagger.tag(temp_left))
                    right_tag=(tagger.tag(temp_right))
                    print(left_tag)
                    print(len(left_tag))
                    for k in range(0,len(left_tag)):
                        tag_l=left_tag[k]
                        print(tag_l)
                        if tag_l[1] == 'PROPN':
                            prop_n_l.append(tag_l[0])
                    for l in range(0,len(right_tag)):
                        tag_r=right_tag[l]
                        print(tag_r)
                        if tag_r[1] == 'PROPN':
                            prop_n_r.append(tag_r[0])
                            prop_n_r
                    corp_l_acquisition.append(prop_n_r)#beta
                    corp_r_acquisition.append(prop_n_l)#baap
                    prop_n_l=[]
                    acquisition_title.append(corpus_train[i])
                    prop_n_r=[]
                    acquisition_id.append(i)
                    check_acquisition.append(id_train[i])
#aakhri acquistion
                #pehla baap
except:
    print(i)
    print(corpus_train[i])
            
            

"""
import re
tree = parser.parse(text)
print('\n'.join(['\t'.join(node) for node in tree]))

w=re.sub('[^a-zA-Z]', ' ',corpus_train[10])
temp='to buy Mines Ltd for $462 mln'
temp
text=temp.split(' ')
tree = parser.parse(text)
print('\n'.join(['\t'.join(node) for node in tree]))
"""

"----------------------------BUY-------------------------------------------------------------------"            

# Writing to an excel  
from xlwt import Workbook 
  
# Workbook is created 
wb = Workbook() 
  
# add_sheet is used to create sheet. 
sheet1 = wb.add_sheet('Sheet 1') 

sheet1.write(0, 0, 'ID')
sheet1.write(0, 1, 'UNIQUE ID')
sheet1.write(0, 2, 'TITLE')
sheet1.write(0, 3, 'ACQ')
sheet1.write(0, 4, 'TAR')
 
for i in range(0, len(buy_id)):
    sheet1.write(i+1, 0, buy_id[i])
    sheet1.write(i+1, 1, check_buy[i])
    sheet1.write(i+1, 2, buy_title[i])
    sheet1.write(i+1, 3, corp_l[i])#baap
    sheet1.write(i+1, 4, corp_r[i])#beta

wb.save('buy.xls')

"----------------------------ACQUIRE--------------------------------------------------------------------"            


from xlwt import Workbook 
  
# Workbook is created 
wb = Workbook() 
  
# add_sheet is used to create sheet. 
sheet1 = wb.add_sheet('Sheet 1') 

sheet1.write(0, 0, 'ID')
sheet1.write(0, 1, 'UNIQUE ID')
sheet1.write(0, 2, 'TITLE')
sheet1.write(0, 3, 'ACQ')
sheet1.write(0, 4, 'TAR')
 
for i in range(0, len(acq_id)):
    sheet1.write(i+1, 0, acq_id[i])
    sheet1.write(i+1, 1, check_acq[i])
    sheet1.write(i+1, 2, acq_title[i])
    sheet1.write(i+1, 3, corp_l_acq[i])#baap
    sheet1.write(i+1, 4, corp_r_acq[i])#beta

wb.save('acquire.xls')


"------------------------------------acquisition-----------------------------------------------------------"            

# Writing to an excel  
from xlwt import Workbook 
  
# Workbook is created 
wb = Workbook() 
  
# add_sheet is used to create sheet. 
sheet1 = wb.add_sheet('Sheet 1') 

sheet1.write(0, 0, 'ID')
sheet1.write(0, 1, 'UNIQUE ID')
sheet1.write(0, 2, 'TITLE')
sheet1.write(0, 3, 'ACQ')
sheet1.write(0, 4, 'TAR')
 
for i in range(0, len(acquisition_id)):
    sheet1.write(i+1, 0, acquisition_id[i])
    sheet1.write(i+1, 1, check_acquisition[i])
    sheet1.write(i+1, 2, acquisition_title[i])
    sheet1.write(i+1, 3, corp_l_acquisition[i])#baap
    sheet1.write(i+1, 4, corp_r_acquisition[i])#beta

wb.save('acquisition.xls')



#TESTING
import xlrd
import numpy as np

a=[]
W = xlrd.open_workbook('MnA_Test.xlsx')
p = W.sheet_by_index(0)
for i in range(1 , 254):
    a.append(p.cell_value(i,0))

from xlwt import Workbook 
  
# Workbook is created 
wb = Workbook() 
  
  
# add_sheet is used to create sheet. 
sheet1 = wb.add_sheet('Sheet 1') 

sheet1.write(0, 0, 'UNIQUE ID')
sheet1.write(0, 1, 'TITLE')
sheet1.write(0, 2, 'ACQ')
sheet1.write(0, 3, 'TAR')
 
for i in range(0, len(acq_id)):
    sheet1.write(i+1, 0, check_acq[i])
    sheet1.write(i+1, 1, acq_title[i])
    sheet1.write(i+1, 2, corp_l_acq[i])#baap
    sheet1.write(i+1, 3, corp_r_acq[i])#beta

q=0
for i in range(38,38+len(buy_id)):
    sheet1.write(i+1, 0, check_buy[q])
    sheet1.write(i+1, 1, buy_title[q])
    sheet1.write(i+1, 2, corp_l[q])#baap
    sheet1.write(i+1, 3, corp_r[q])#beta
    q=q+1

v=0
for i in range(100,100+len(acquisition_id)):
    sheet1.write(i+1, 0, check_acquisition[v])
    sheet1.write(i+1, 1, acquisition_title[v])
    sheet1.write(i+1, 2, corp_l_acquisition[v])#baap
    sheet1.write(i+1, 3, corp_r_acquisition[v])#beta
    v=v+1

    
wb.save('test1.xls')



